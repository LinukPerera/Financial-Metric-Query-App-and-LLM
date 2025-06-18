import openpyxl
from openpyxl.utils import get_column_letter

def get_merged_parent(sheet, row, col):
    """Return the merged value and its full bounds if cell is in a merged range."""
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row <= row <= max_row and min_col <= col <= max_col:
            value = sheet.cell(row=min_row, column=min_col).value
            return value, (min_row, max_row), (min_col, max_col)
    value = sheet.cell(row=row, column=col).value
    return value, (row, row), (col, col)

def detect_first_nonempty_row(sheet, max_scan=10):
    for i in range(1, max_scan + 1):
        if any(cell.value not in (None, '') for cell in sheet[i]):
            return i
    return 1

def count_filled_header_cells(sheet, start_row, num_rows=2):
    filled = set()
    for row in range(start_row, start_row + num_rows):
        if row > sheet.max_row:
            continue
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=row, column=col).value
            if val not in (None, ''):
                filled.add((row, col))
    return len(filled)

def detect_multirow_sheet(wb):
    counts = {}
    for name in wb.sheetnames[:3]:
        sheet = wb[name]
        start = detect_first_nonempty_row(sheet)
        counts[name] = count_filled_header_cells(sheet, start, num_rows=2)
    return max(counts, key=counts.get)

def extract_header_signature(sheet, header_rows):
    """Return a tuple representing the text content of the header for comparison."""
    signature = []
    for col in range(1, sheet.max_column + 1):
        labels = []
        for row in header_rows:
            value = sheet.cell(row=row, column=col).value
            if value not in (None, ''):
                labels.append(str(value).strip())
        signature.append(tuple(labels))
    return tuple(signature)

def find_repeating_headers(sheet, base_header_rows):
    """Scan the sheet and return rows where the same header appears again."""
    base_signature = extract_header_signature(sheet, base_header_rows)
    repeat_rows = []

    for row in range(base_header_rows[-1] + 1, sheet.max_row - 1):
        next_rows = [row + i for i in range(len(base_header_rows))]
        test_signature = extract_header_signature(sheet, next_rows)
        if test_signature == base_signature:
            repeat_rows.append(row)

    return repeat_rows

def detect_sectors(sheet, glossary_start=None):
    """Detect rows where column A has >4 characters and column B is blank.
       Stops checking after the glossary starts (if defined)."""
    sectors = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        row_num = row[0].row
        if glossary_start and row_num >= glossary_start:
            break  # Stop checking after glossary

        code = row[0].value
        b_val = row[1].value if len(row) > 1 else None
        if isinstance(code, str) and len(code.strip()) > 4 and (b_val in (None, '')):
            sectors.append((row_num, code.strip()))

    if sectors:
        print("\n🏷️ Sector identifiers found:")
        for row_num, code in sectors:
            print(f"Row {row_num}: Sector Code = {code}")
            # Print first 30 non-blank cells from that row
            row_cells = [sheet.cell(row=row_num, column=col).value for col in range(1, 31)]
            non_blank_cells = [str(cell).strip() for cell in row_cells if cell not in (None, '')]
            if non_blank_cells:
                print("  ➤ Row content (first 30 cells, non-blank):", ' | '.join(non_blank_cells))
    return sectors



def detect_glossary(sheet, header_rows, repeating_header_rows, sector_rows):
    header_start_col = 1
    header_end_col = sheet.max_column

    exclude_rows = set(repeating_header_rows)
    exclude_rows.update(row + 1 for row in repeating_header_rows)
    exclude_rows.update(sector_rows)

    for row in range(header_rows[-1] + 1, sheet.max_row + 1):
        if row in exclude_rows:
            continue

        filled_count = 0
        colon_found = False
        for col in range(header_start_col, header_end_col + 1):
            val = sheet.cell(row=row, column=col).value
            if isinstance(val, str) and ':' in val:
                colon_found = True
            if val not in (None, ''):
                filled_count += 1

        # Check for partial fill and colon
        if colon_found and 0 < filled_count < (header_end_col - header_start_col + 1):
            print(f"\n📘 Glossary starts at row {row}")
            for r in range(row, min(sheet.max_row + 1, row + 6)):
                row_values = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
                if any(cell not in (None, '') for cell in row_values):
                    text_line = ' | '.join(str(cell).strip() for cell in row_values if cell not in (None, ''))
                    print(f"Row {r}: {text_line}")
            print(f"\n🔚 Beyond row {row}, this is considered the glossary.")
            return row
    return None

def print_sheet_headers(sheet, multirow_enabled):
    print(f"\n--- Headers from sheet: {sheet.title} ---")

    base_row = detect_first_nonempty_row(sheet)
    header_rows = [base_row]
    if multirow_enabled:
        header_rows.append(base_row + 1)

    # 1. Print header hierarchy
    for col in range(1, sheet.max_column + 1):
        hierarchy = []
        for row in header_rows:
            if row > sheet.max_row:
                continue
            value, (row_start, row_end), (col_start, col_end) = get_merged_parent(sheet, row, col)
            if value not in (None, ''):
                label = str(value).strip()
                if label not in hierarchy:
                    hierarchy.append(label)

        col_letter = get_column_letter(col)
        if hierarchy:
            print(f"{col_letter}: {' > '.join(hierarchy)}")
        else:
            print(f"{col_letter}: [Blank]")

    # 2. Detect repeating headers
    repeating = find_repeating_headers(sheet, header_rows)
    if repeating:
        print(f"\n🔁 Repeating header rows found at: {repeating}")
    else:
        print("\n✅ No repeating headers detected.")

    # 3. Detect glossary
    sector_rows = []
    glossary_start = detect_glossary(sheet, header_rows, repeating, sector_rows)

    # 4. Detect sectors only *after* glossary detection
    detect_sectors(sheet, glossary_start)

def main():
    file_path = "financial_data.xlsx" 
    wb = openpyxl.load_workbook(file_path, data_only=True)

    multirow_sheet = detect_multirow_sheet(wb)

    for sheet_name in wb.sheetnames[:3]:
        sheet = wb[sheet_name]
        is_multirow = (sheet_name == multirow_sheet)
        print_sheet_headers(sheet, multirow_enabled=is_multirow)

if __name__ == "__main__":
    main()
