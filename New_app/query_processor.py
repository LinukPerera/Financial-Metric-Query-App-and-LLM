import openpyxl
from openpyxl.utils import get_column_letter
import re
import sys
import logging
from contextlib import contextmanager
import plotly.graph_objects as go
from DebugHeader import (
    detect_first_nonempty_row,
    detect_multirow_sheet,
    extract_header_signature,
    find_repeating_headers,
    detect_sectors,
    detect_glossary,
    get_merged_parent,
)
from metric_dictionary import metric_mappings, metric_definitions

# Set up logging
logging.basicConfig(filename='debug.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

@contextmanager
def suppress_prints():
    """Suppress print statements from DebugHeader.py."""
    original_stdout = sys.stdout
    sys.stdout = open('/dev/null', 'w')
    try:
        yield
    finally:
        sys.stdout.close()
        sys.stdout = original_stdout

class FinancialQueryProcessor:
    def __init__(self, file_path):
        with suppress_prints():
            try:
                self.wb = openpyxl.load_workbook(file_path, data_only=True)
                self.sheet_info = self._parse_sheets()
                self.primary_sheet = self.sheet_info[detect_multirow_sheet(self.wb)]
                logging.info(f"Initialized with primary sheet: {self.primary_sheet['sheet'].title}")
                logging.debug(f"Sheet info keys: {list(self.sheet_info.keys())}")
            except Exception as e:
                logging.error(f"Failed to initialize workbook: {e}")
                raise

    def _parse_sheets(self):
        """Parse all sheets to extract headers, sectors, and data ranges."""
        sheet_info = {}
        with suppress_prints():
            for sheet_name in self.wb.sheetnames:
                sheet = self.wb[sheet_name]
                header_start = detect_first_nonempty_row(sheet)
                is_multirow = (sheet_name == detect_multirow_sheet(self.wb))
                header_rows = [header_start] if not is_multirow else [header_start, header_start + 1]
                logging.debug(f"Parsing sheet: {sheet_name}, header_rows: {header_rows}")

                headers = {}
                code_columns = {}
                for col in range(1, sheet.max_column + 1):
                    hierarchy = []
                    has_code = False
                    for row in header_rows:
                        if row > sheet.max_row:
                            continue
                        value, _, _ = get_merged_parent(sheet, row, col)
                        if value not in (None, ''):
                            hierarchy.append(str(value).strip())
                            if str(value).strip().upper() == 'CODE':
                                has_code = True
                                code_columns[col] = str(value).strip()
                    headers[get_column_letter(col)] = ' > '.join(hierarchy) if hierarchy else None
                    if has_code:
                        code_columns[col] = code_columns.get(col, str(value).strip())

                code_count = len(code_columns)
                primary_key_col = 'B' if (code_count == 2 or (code_count == 1 and 2 in code_columns)) else 'A'
                logging.debug(f"Sheet {sheet_name}: code_columns={code_columns}, code_count={code_count}, primary_key_col={primary_key_col}")

                repeating_headers = find_repeating_headers(sheet, header_rows)
                glossary_start = detect_glossary(sheet, header_rows, repeating_headers, [])
                sectors = detect_sectors(sheet, glossary_start)
                sector_rows = [row for row, _ in sectors]
                logging.debug(f"Sectors: {sectors}, Repeating headers: {repeating_headers}")

                exclude_rows = set(header_rows + repeating_headers + sector_rows)
                if glossary_start:
                    exclude_rows.update(range(glossary_start, sheet.max_row + 1))
                data_rows = [r for r in range(header_rows[-1] + 1, sheet.max_row + 1) if r not in exclude_rows]
                logging.debug(f"Data rows: {data_rows[:5]}... (total {len(data_rows)})")

                sheet_info[sheet_name] = {
                    'sheet': sheet,
                    'headers': headers,
                    'header_rows': header_rows,
                    'sectors': sectors,
                    'data_rows': data_rows,
                    'primary_key_col': primary_key_col,
                }
        return sheet_info
    
    def _normalize_metric(self, metric):
        """Normalize metric name using metric_mappings."""
        if not metric:
            return metric
        metric_clean = metric.lower().replace('.', '').replace('/', '').strip()
        for raw_metric, mapped_metric in metric_mappings.items():
            raw_clean = raw_metric.lower().replace('.', '').replace('/', '')
            mapped_clean = mapped_metric.lower().replace('.', '').replace('/', '')
            if metric_clean == raw_clean or metric_clean == mapped_clean:
                return mapped_metric
        return metric

    def _find_column_across_sheets(self, keyword):
        """Find columns matching a keyword across all sheets, handling multi-word metrics."""
        results = []
        # Try different normalization approaches
        search_terms = [
            keyword,                            # Original
            keyword.replace(' ', ''),           # Without spaces
            keyword.replace(' ', '.'),          # With dots instead of spaces
            keyword.split(' ')[0]               # First word only
        ]
        
        for sheet_name in self.sheet_info:
            headers = self.sheet_info[sheet_name]['headers']
            pe_columns = []  # Track P/E columns for this sheet
            
            for col, header in headers.items():
                if header:
                    # Split header at '>' to isolate main metric
                    main_metric = header.split(' > ')[0] if ' > ' in header else header
                    # Normalize for comparison
                    normalized_header = main_metric.lower().replace(' ', '')
                    
                    # Check against all possible search term variations
                    for term in search_terms:
                        normalized_term = term.lower().replace(' ', '')
                        if normalized_term == normalized_header:
                            if normalized_term == 'pe':
                                pe_columns.append((sheet_name, col, header))
                            else:
                                results.append((sheet_name, col, header))
                            break
            
            # For P/E, select only the earliest column
            if pe_columns:
                earliest_pe = min(pe_columns, key=lambda x: openpyxl.utils.column_index_from_string(x[1]))
                results.append(earliest_pe)
        
        logging.debug(f"Finding '{keyword}' across sheets: {results}")
        return results

    def _find_company(self, sheet_name, code):
        """Find a company's data row by its exact code, handling extensions properly."""
        sheet = self.sheet_info[sheet_name]['sheet']
        primary_key_col = self.sheet_info[sheet_name]['primary_key_col']
        col_idx = openpyxl.utils.column_index_from_string(primary_key_col) - 1
        code = str(code).strip()
        
        # First try exact match (case-sensitive)
        for row in self.sheet_info[sheet_name]['data_rows']:
            cell_value = sheet.cell(row=row, column=col_idx + 1).value
            if cell_value and str(cell_value).strip() == code:
                return row
        
        # If code has an extension, try with just the base code
        if '.' in code:
            base_code = code.split('.')[0]
            for row in self.sheet_info[sheet_name]['data_rows']:
                cell_value = sheet.cell(row=row, column=col_idx + 1).value
                if cell_value and str(cell_value).strip().upper() == base_code.upper():
                    return row
        
        return None

    def _get_sector_rows(self, sheet_name, sector_code):
        """Get all data rows for a given sector, case-insensitive."""
        sectors = self.sheet_info[sheet_name]['sectors']
        data_rows = self.sheet_info[sheet_name]['data_rows']
        sector_code = sector_code.upper().strip()
        for i, (sector_row, sector_name) in enumerate(sectors):
            if sector_name.upper() == sector_code:
                next_sector_row = sectors[i + 1][0] if i + 1 < len(sectors) else self.sheet_info[sheet_name]['sheet'].max_row + 1
                rows = [r for r in data_rows if sector_row < r < next_sector_row]
                logging.debug(f"Sector '{sector_code}' rows: {rows}")
                return rows
        logging.debug(f"Sector '{sector_code}' not found in {sheet_name}")
        return []
    
    def _get_all_sectors(self, sheet_name):
        """Get all sector names in a sheet."""
        sectors = self.sheet_info[sheet_name]['sectors']
        sector_names = [sector_name for _, sector_name in sectors]
        logging.debug(f"All sectors in {sheet_name}: {sector_names}")
        return sector_names

    def _define_header(self, metric):
        """Provide a definition for a financial metric using metric_definitions."""
        metric = metric_mappings.get(metric.lower().replace('.', '').strip(), metric.lower().replace('.', '').strip())
        definition = metric_definitions.get(metric, None)
        if definition:
            return definition, None
        matches = self._find_column_across_sheets(metric.lower())
        if matches:
            sheets = set(sheet_name for sheet_name, _, _ in matches)
            return f"'{metric}' is a financial metric found in sheets {', '.join(sorted(sheets))}. No detailed definition available.", None
        logging.debug(f"No definition found for '{metric}'")
        return f"Sorry, I don’t have a definition for '{metric}'. Try metrics like 'P/E', 'Div Yield', or 'Revenue 3M'.", None

    def _handle_company_multi_metric(self, code, metrics):
        """Retrieve multiple metrics for a company, handling extensions properly."""
        results = []
        chart_data = None
        row_cache = {}
        original_code = code.upper().strip()
        
        for metric in metrics:
            matches = self._find_column_across_sheets(metric)
            if not matches:
                results.append(f"Metric '{metric}' not found in any sheet.")
                continue
            
            for sheet_name, col, header in matches:
                if sheet_name not in row_cache:
                    # First try exact match
                    row_cache[sheet_name] = self._find_company(sheet_name, original_code)
                    
                    # If not found and has extension, try base code
                    if not row_cache[sheet_name] and '.' in original_code:
                        base_code = original_code.split('.')[0]
                        row_cache[sheet_name] = self._find_company(sheet_name, base_code)
                
                row = row_cache[sheet_name]
                if not row:
                    results.append(f"Company {original_code} not found in {sheet_name} for metric '{metric}'.")
                    continue
                
                sheet = self.sheet_info[sheet_name]['sheet']
                value = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                display_header = header.split(' > ')[-1] if ' > ' in header else header
                results.append(f"{sheet_name} - {display_header}: {value if value is not None else 'N/A'}")
        
        if not results:
            result = f"No data found for {original_code} with metrics {', '.join(metrics)}."
        else:
            result = f"Data for {original_code}:\n" + "\n".join(results)
            chart_data = self._generate_company_chart(original_code, metrics)
        
        return result, chart_data
    
    #### New section updated 0n 17/10/2025

    def _handle_sector_metric(self, sector, metric):
        """Retrieve a metric for a sector with comparisons - display average and all companies."""
        sheet_name = detect_multirow_sheet(self.wb)
        columns = [col for sheet, col, _ in self._find_column_across_sheets(metric) if sheet == sheet_name]
        if not columns:
            return f"Metric '{metric}' not found in {sheet_name}.", None
        
        if sector.lower().strip() == 'all sectors':
            # Handle "all sectors" case - show all sectors for the metric
            all_sectors = self._get_all_sectors(sheet_name)
            results = []
            sector_data = []
            
            for sector_name in all_sectors:
                sector_result, avg_value = self._compute_sector_average(sheet_name, sector_name, metric, columns)
                if sector_result:
                    results.append(sector_result)
                    sector_data.append((sector_name, avg_value))
            
            if not results:
                return f"No data found for {metric} across sectors.", None
            
            # Format results for all sectors
            result = f"{metric} Across All Sectors in {sheet_name}:\n\n"
            for sector_name, avg_value in sector_data:
                result += f"• {sector_name}: Average {metric} = {avg_value:.2f}\n"
            
            # Generate chart comparing all sectors
            chart_data = self._generate_all_sectors_chart(sector_data, metric)
            return result, chart_data
        else:
            # Handle single sector case - show average and all companies
            result, chart_data = self._compute_sector_metric_with_companies(sheet_name, sector, metric, columns)
            if not result:
                return f"Sector '{sector}' or data for '{metric}' not found in {sheet_name}.", None
            return result, chart_data

    def _compute_sector_average(self, sheet_name, sector, metric, columns):
        """Compute average metric value for a sector."""
        rows = self._get_sector_rows(sheet_name, sector)
        if not rows:
            return None, None
        
        sheet = self.sheet_info[sheet_name]['sheet']
        values = []
        
        for row in rows:
            for col in columns:
                value = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                if isinstance(value, (int, float)):
                    values.append(value)
        
        if not values:
            return None, None
        
        avg_value = sum(values) / len(values)
        return f"{sector}: {avg_value:.2f}", avg_value

    def _compute_sector_metric_with_companies(self, sheet_name, sector, metric, columns):
        """Compute metric for a sector and show all companies with their values."""
        rows = self._get_sector_rows(sheet_name, sector)
        if not rows:
            return None, None
        
        sheet = self.sheet_info[sheet_name]['sheet']
        primary_key_col = self.sheet_info[sheet_name]['primary_key_col']
        col_idx = openpyxl.utils.column_index_from_string(primary_key_col) - 1
        values = []
        company_data = []
        
        for row in rows:
            code = sheet.cell(row=row, column=col_idx + 1).value
            if not code:
                continue
            for col in columns:
                value = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                if isinstance(value, (int, float)):
                    values.append(value)
                    company_data.append((code.strip(), value))
        
        if not values:
            return None, None
        
        avg_value = sum(values) / len(values)
        
        # Format result with average and all companies
        result = (
            f"{metric} for sector {sector} in {sheet_name}:\n"
            f"Average: {avg_value:.2f}\n\n"
            f"Companies in {sector}:\n"
        )
        
        # Sort companies by metric value (highest to lowest)
        company_data.sort(key=lambda x: x[1], reverse=True)
        for code, value in company_data:
            result += f"  • {code}: {value:.2f}\n"
        
        chart_data = self._generate_sector_chart(sector, metric, company_data)
        logging.debug(f"Sector metric result for {sector}: {result}")
        return result, chart_data

    def _generate_all_sectors_chart(self, sector_data, metric):
        """Generate a bar chart comparing all sectors for a metric."""
        if not sector_data:
            return None
        
        # Sort sectors by average value (highest to lowest)
        sector_data.sort(key=lambda x: x[1], reverse=True)
        sectors, values = zip(*sector_data)
        
        fig = go.Figure(data=[
            go.Bar(x=sectors, y=values, marker_color='#FF6384', name=f'Average {metric}')
        ])
        
        fig.update_layout(
            title=f'Average {metric} Across All Sectors',
            xaxis_title='Sector',
            yaxis_title=f'Average {metric}',
            template='plotly_white',
            xaxis_tickangle=-45
        )
        
        return fig

    def _handle_general_metric(self, metric):
        """Retrieve a metric across all data in the primary sheet with more detailed info."""
        sheet_name = detect_multirow_sheet(self.wb)
        columns = [col for sheet, col, _ in self._find_column_across_sheets(metric) if sheet == sheet_name]
        if not columns:
            return f"Metric '{metric}' not found in {sheet_name}.", None
        
        sheet = self.sheet_info[sheet_name]['sheet']
        values = []
        company_data = []
        primary_key_col = self.sheet_info[sheet_name]['primary_key_col']
        col_idx = openpyxl.utils.column_index_from_string(primary_key_col) - 1
        
        for row in self.sheet_info[sheet_name]['data_rows']:
            code = sheet.cell(row=row, column=col_idx + 1).value
            if not code:
                continue
            for col in columns:
                value = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                if isinstance(value, (int, float)):
                    values.append(value)
                    company_data.append((code.strip(), value))
        
        if not values:
            return f"No numerical data found for {metric} in {sheet_name}.", None
        
        avg_value = sum(values) / len(values)
        min_value = min(values)
        max_value = max(values)
        
        # Get top 5 companies with highest values
        top_companies = sorted(company_data, key=lambda x: x[1], reverse=True)[:5]
        
        result = (
            f"Analysis of {metric} across all companies in {sheet_name}:\n"
            f"Average: {avg_value:.2f}\n"
            f"Range: {min_value:.2f} - {max_value:.2f}\n"
            f"Total Companies: {len(values)}\n\n"
            f"Top 5 Companies by {metric}:\n"
        )
        
        for idx, (code, value) in enumerate(top_companies, 1):
            result += f"  {idx}. {code}: {value:.2f}\n"
        
        logging.debug(f"General metric result: {result}")
        return result, None

    def _find_best_stock(self, criteria):
        """Find top stocks by a metric or composite score."""
        sheet_name = detect_multirow_sheet(self.wb)
        sheet = self.sheet_info[sheet_name]['sheet']
        data_rows = self.sheet_info[sheet_name]['data_rows']
        primary_key_col = self.sheet_info[sheet_name]['primary_key_col']
        col_idx = openpyxl.utils.column_index_from_string(primary_key_col) - 1
        if not criteria:
            metrics = ['P/E', 'Div Yield', 'PBV']
            weights = {'P/E': 0.4, 'Div Yield': 0.3, 'PBV': -0.2}
        else:
            metrics = [criteria]
            weights = {criteria: 1.0}
        scores = []
        for row in data_rows:
            code = sheet.cell(row=row, column=col_idx + 1).value
            if not code:
                continue
            score = 0
            valid = True
            for metric in metrics:
                columns = [col for sheet, col, _ in self._find_column_across_sheets(metric) if sheet == sheet_name]
                if not columns:
                    return f"Metric '{metric}' not found in {sheet_name}.", None
                value = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(columns[0])).value
                if not isinstance(value, (int, float)):
                    valid = False
                    break
                score += value * weights[metric]
            if valid:
                scores.append((code.strip(), score))
        if not scores:
            return "No valid data found to rank stocks.", None
        sorted_scores = sorted(scores, key=lambda x: x[1], reverse=True)[:3]
        result = f"Top stocks by {criteria or 'composite score (P/E, Div Yield, PBV)'} in {sheet_name}:\n"
        for idx, (code, score) in enumerate(sorted_scores):
            result += f"{idx + 1}. {code}: Score = {score:.2f}\n"
        chart_data = self._generate_best_stock_chart(sorted_scores, criteria or 'Composite Score')
        logging.debug(f"Best stock result: {result}")
        return result, chart_data

    def _find_best_sector(self, criteria):
        """Find top sectors by average metric value."""
        sheet_name = detect_multirow_sheet(self.wb)
        all_sectors = self._get_all_sectors(sheet_name)
        if not criteria:
            return "Please specify a metric (e.g., 'P/E') to rank sectors.", None
        columns = [col for sheet, col, _ in self._find_column_across_sheets(criteria) if sheet == sheet_name]
        if not columns:
            return f"Metric '{criteria}' not found in {sheet_name}.", None
        sheet = self.sheet_info[sheet_name]['sheet']
        sector_scores = []
        for sector in all_sectors:
            rows = self._get_sector_rows(sheet_name, sector)
            values = []
            for row in rows:
                for col in columns:
                    value = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                    if isinstance(value, (int, float)):
                        values.append(value)
            if values:
                avg_value = sum(values) / len(values)
                sector_scores.append((sector, avg_value))
        if not sector_scores:
            return f"No valid data found for {criteria} across sectors.", None
        sorted_sectors = sorted(sector_scores, key=lambda x: x[1], reverse=True)[:3]
        result = f"Top sectors by {criteria} in {sheet_name}:\n"
        for idx, (sector, score) in enumerate(sorted_sectors):
            result += f"{idx + 1}. {sector}: Average {criteria} = {score:.2f}\n"
        chart_data = self._generate_sector_comparison_chart(sorted_sectors, criteria)
        logging.debug(f"Best sector result: {result}")
        return result, chart_data

    def _find_lowest_metric_value(self, metric):
        """Find companies with the lowest values for a metric."""
        sheet_name = detect_multirow_sheet(self.wb)
        
        # Find columns for the metric
        columns = [col for sheet, col, _ in self._find_column_across_sheets(metric) if sheet == sheet_name]
        
        # If not found, try alternative representations
        if not columns:
            alt_metric = metric.replace(' ', '')
            columns = [col for sheet, col, _ in self._find_column_across_sheets(alt_metric) if sheet == sheet_name]
            
            if not columns and ' ' in metric:
                first_word = metric.split(' ')[0]
                columns = [col for sheet, col, _ in self._find_column_across_sheets(first_word) if sheet == sheet_name]
        
        if not columns:
            return f"Metric '{metric}' not found in {sheet_name}.", None
        
        sheet = self.sheet_info[sheet_name]['sheet']
        primary_key_col = self.sheet_info[sheet_name]['primary_key_col']
        col_idx = openpyxl.utils.column_index_from_string(primary_key_col) - 1
        data_rows = self.sheet_info[sheet_name]['data_rows']
        metric_values = []
        
        # Collect all metric values
        for row in data_rows:
            code = sheet.cell(row=row, column=col_idx + 1).value
            if not code:
                continue
            for col in columns:
                value = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                if isinstance(value, (int, float)):
                    metric_values.append((code.strip(), value))
        
        if not metric_values:
            return f"No valid data found for {metric} in {sheet_name}.", None
        
        # Sort by lowest values (ascending)
        sorted_values = sorted(metric_values, key=lambda x: x[1])[:3]
        
        result = f"Top companies with lowest {metric} in {sheet_name}:\n"
        for idx, (code, value) in enumerate(sorted_values):
            result += f"{idx + 1}. {code}: {metric} = {value:.2f}\n"
        
        chart_data = self._generate_best_metric_chart(sorted_values, metric, 'min')
        logging.debug(f"Lowest metric result: {result}")
        return result, chart_data

    def _find_highest_metric_value(self, metric):
        """Find companies with the highest values for a metric."""
        sheet_name = detect_multirow_sheet(self.wb)
        
        # Find columns for the metric
        columns = [col for sheet, col, _ in self._find_column_across_sheets(metric) if sheet == sheet_name]
        
        # If not found, try alternative representations
        if not columns:
            alt_metric = metric.replace(' ', '')
            columns = [col for sheet, col, _ in self._find_column_across_sheets(alt_metric) if sheet == sheet_name]
            
            if not columns and ' ' in metric:
                first_word = metric.split(' ')[0]
                columns = [col for sheet, col, _ in self._find_column_across_sheets(first_word) if sheet == sheet_name]
        
        if not columns:
            return f"Metric '{metric}' not found in {sheet_name}.", None
        
        sheet = self.sheet_info[sheet_name]['sheet']
        primary_key_col = self.sheet_info[sheet_name]['primary_key_col']
        col_idx = openpyxl.utils.column_index_from_string(primary_key_col) - 1
        data_rows = self.sheet_info[sheet_name]['data_rows']
        metric_values = []
        
        # Collect all metric values
        for row in data_rows:
            code = sheet.cell(row=row, column=col_idx + 1).value
            if not code:
                continue
            for col in columns:
                value = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                if isinstance(value, (int, float)):
                    metric_values.append((code.strip(), value))
        
        if not metric_values:
            return f"No valid data found for {metric} in {sheet_name}.", None
        
        # Sort by highest values (descending)
        sorted_values = sorted(metric_values, key=lambda x: x[1], reverse=True)[:3]
        
        result = f"Top companies with highest {metric} in {sheet_name}:\n"
        for idx, (code, value) in enumerate(sorted_values):
            result += f"{idx + 1}. {code}: {metric} = {value:.2f}\n"
        
        chart_data = self._generate_best_metric_chart(sorted_values, metric, 'max')
        logging.debug(f"Highest metric result: {result}")
        return result, chart_data

    def _find_best_metric_value(self, metric):
        """Find and display both highest and lowest values for a metric."""
        sheet_name = detect_multirow_sheet(self.wb)
        
        # Find columns for the metric
        columns = [col for sheet, col, _ in self._find_column_across_sheets(metric) if sheet == sheet_name]
        
        # If not found, try alternative representations
        if not columns:
            alt_metric = metric.replace(' ', '')
            columns = [col for sheet, col, _ in self._find_column_across_sheets(alt_metric) if sheet == sheet_name]
            
            if not columns and ' ' in metric:
                first_word = metric.split(' ')[0]
                columns = [col for sheet, col, _ in self._find_column_across_sheets(first_word) if sheet == sheet_name]
        
        if not columns:
            return f"Metric '{metric}' not found in {sheet_name}.", None
        
        sheet = self.sheet_info[sheet_name]['sheet']
        primary_key_col = self.sheet_info[sheet_name]['primary_key_col']
        col_idx = openpyxl.utils.column_index_from_string(primary_key_col) - 1
        data_rows = self.sheet_info[sheet_name]['data_rows']
        metric_values = []
        
        # Collect all metric values
        for row in data_rows:
            code = sheet.cell(row=row, column=col_idx + 1).value
            if not code:
                continue
            for col in columns:
                value = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                if isinstance(value, (int, float)):
                    metric_values.append((code.strip(), value))
        
        if not metric_values:
            return f"No valid data found for {metric} in {sheet_name}.", None
        
        # Determine if lower or higher is typically better for this metric
        lower_is_better = [
            'pe', 'pbv', 'issuedqtymn', 'issued', 'qty', 'mn'
        ]
        
        normalized_metric = self._normalize_metric(metric).lower().replace(' ', '').replace('.', '').replace('/', '')
        is_lower_better = normalized_metric in lower_is_better
        
        # Get top 3 highest values
        highest_values = sorted(metric_values, key=lambda x: x[1], reverse=True)[:3]
        
        # Get top 3 lowest values  
        lowest_values = sorted(metric_values, key=lambda x: x[1])[:3]
        
        # Build comprehensive result
        result = f"Analysis of {metric} in {sheet_name}:\n\n"
        
        if is_lower_better:
            result += f"🔽 BEST (Lowest {metric} - Typically Preferred):\n"
            for idx, (code, value) in enumerate(lowest_values):
                result += f"  {idx + 1}. {code}: {value:.2f}\n"
            
            result += f"\n🔼 HIGHEST {metric}:\n"
            for idx, (code, value) in enumerate(highest_values):
                result += f"  {idx + 1}. {code}: {value:.2f}\n"
                
            result += f"\n💡 For {metric}, lower values are typically better."
            
        else:
            result += f"🔼 BEST (Highest {metric} - Typically Preferred):\n"
            for idx, (code, value) in enumerate(highest_values):
                result += f"  {idx + 1}. {code}: {value:.2f}\n"
            
            result += f"\n🔽 LOWEST {metric}:\n"
            for idx, (code, value) in enumerate(lowest_values):
                result += f"  {idx + 1}. {code}: {value:.2f}\n"
                
            result += f"\n💡 For {metric}, higher values are typically better."
        
        # Generate chart showing both extremes
        chart_data = self._generate_best_metric_comparison_chart(highest_values, lowest_values, metric, is_lower_better)
        
        logging.debug(f"Best metric analysis - Lower is better: {is_lower_better}")
        return result, chart_data

    def _generate_best_metric_comparison_chart(self, highest_values, lowest_values, metric, is_lower_better):
        """Generate a comparison chart showing both highest and lowest values."""
        if not highest_values or not lowest_values:
            return None
        
        # Combine and label the data
        all_companies = []
        all_values = []
        colors = []
        
        # Add highest values
        for code, value in highest_values:
            all_companies.append(f"{code} (Highest)")
            all_values.append(value)
            colors.append('#FF6B6B' if is_lower_better else '#4ECDC4')  # Red if lower is better, green if higher is better
        
        # Add lowest values  
        for code, value in lowest_values:
            all_companies.append(f"{code} (Lowest)")
            all_values.append(value)
            colors.append('#4ECDC4' if is_lower_better else '#FF6B6B')  # Green if lower is better, red if higher is better
        
        fig = go.Figure(data=[
            go.Bar(x=all_companies, y=all_values, marker_color=colors, name=metric)
        ])
        
        preferred_direction = "Lower" if is_lower_better else "Higher"
        fig.update_layout(
            title=f'{metric} - Highest vs Lowest Values (Best: {preferred_direction})',
            xaxis_title='Company',
            yaxis_title=metric,
            template='plotly_white',
            showlegend=False
        )
        
        return fig
       
    def _compare_stocks(self, stocks, metric='P/E'):
        """Compare stocks while properly handling extensions."""
        sheet_name = detect_multirow_sheet(self.wb)
        columns = [col for sheet, col, header in self._find_column_across_sheets(metric) 
                if sheet == sheet_name and '%' not in header]
        
        if not columns:
            return f"Metric '{metric}' not found in {sheet_name}.", None
        
        results = []
        chart_data = []
        
        for original_code in stocks:
            code = original_code.strip().upper()
            row = self._find_company(sheet_name, code)
            
            if not row and '.' in code:
                # Try with base code if extension version not found
                base_code = code.split('.')[0]
                row = self._find_company(sheet_name, base_code)
            
            if not row:
                results.append(f"{code}: Company not found in {sheet_name}.")
                continue
                
            sheet = self.sheet_info[sheet_name]['sheet']
            value = None
            for col in columns:
                val = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                if isinstance(val, (int, float)):
                    value = val
                    break
                    
            if value is not None:
                results.append(f"{code}: {metric} = {value:.2f}")
                chart_data.append((code, value))
            else:
                results.append(f"{code}: No valid {metric} data in {sheet_name}.")
        
        if not chart_data:
            return f"No valid data found for {metric} across specified stocks.", None
            
        # Generate comparison chart
        labels, values = zip(*chart_data)
        fig = go.Figure(data=[
            go.Bar(x=labels, y=values, marker_color='#36A2EB', name=metric)
        ])
        fig.update_layout(
            title=f'Comparison of {metric} Across Stocks',
            xaxis_title='Stock',
            yaxis_title=metric,
            template='plotly_white'
        )
        
        return "Comparison of stocks:\n" + "\n".join(results), fig
    

    def _handle_multi_sheet_query(self, code, metric_sheet_pairs):
        """Handle queries specifying metrics from specific sheets."""
        results = []
        chart_data = None
        row_cache = {}
        code = code.upper().strip()
        for metric, sheet_name in metric_sheet_pairs:
            matches = [(s, c, h) for s, c, h in self._find_column_across_sheets(metric) if s.lower() == sheet_name.lower()]
            if not matches:
                results.append(f"Metric '{metric}' not found in sheet '{sheet_name}'.")
                continue
            for sheet_name, col, header in matches:
                if sheet_name not in row_cache:
                    row_cache[sheet_name] = self._find_company(sheet_name, code)
                row = row_cache[sheet_name]
                if not row:
                    results.append(f"Company {code} not found in {sheet_name} for metric '{metric}'.")
                continue
                sheet = self.sheet_info[sheet_name]['sheet']
                value = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                display_header = header.split(' > ')[-1] if ' > ' in header else header
                results.append(f"{sheet_name} - {display_header}: {value if value is not None else 'N/A'}")
        if not results:
            result = f"No data found for {code} with specified metrics."
        else:
            result = f"Data for {code}:\n" + "\n".join(results)
            chart_data = self._generate_multi_sheet_chart(code, metric_sheet_pairs)
        logging.debug(f"Multi-sheet query result: {result}")
        return result, chart_data
    
    def _compare_mixed_entities(self, entities, metric):
        """Compare companies and sectors by a metric."""
        sheet_name = detect_multirow_sheet(self.wb)
        columns = [col for sheet, col, _ in self._find_column_across_sheets(metric) if sheet == sheet_name]
        if not columns:
            return f"Metric '{metric}' not found in {sheet_name}.", None

        results = []
        labels = []
        values = []

        for entity in entities:
            entity_name = entity['name'].upper().strip()
            entity_type = entity['type']
            
            if not self._validate_entity(entity_name, is_company=(entity_type == 'company')):
                results.append(f"{entity_type.capitalize()} '{entity_name}' not found in {sheet_name}.")
                continue
                
            if entity_type == 'company':
                # Use the improved company finding that handles extensions
                row = self._find_company(sheet_name, entity_name)
                if not row:
                    results.append(f"Company '{entity_name}' not found in {sheet_name}.")
                    continue
                    
                sheet = self.sheet_info[sheet_name]['sheet']
                value = None
                for col in columns:
                    val = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                    if isinstance(val, (int, float)):
                        value = val
                        break
                        
                if value is not None:
                    results.append(f"Company {entity_name}: {metric} = {value:.2f}")
                    labels.append(entity_name)
                    values.append(value)
                else:
                    results.append(f"No valid {metric} data for company {entity_name} in {sheet_name}.")
                    
            elif entity_type == 'sector':
                rows = self._get_sector_rows(sheet_name, entity_name)
                if not rows:
                    results.append(f"Sector '{entity_name}' not found in {sheet_name}.")
                    continue
                    
                sheet = self.sheet_info[sheet_name]['sheet']
                sector_values = []
                for row in rows:
                    for col in columns:
                        val = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                        if isinstance(val, (int, float)):
                            sector_values.append(val)
                            
                if sector_values:
                    avg_value = sum(sector_values) / len(sector_values)
                    results.append(f"Sector {entity_name}: Average {metric} = {avg_value:.2f}")
                    labels.append(f"{entity_name} (Sector Avg)")
                    values.append(avg_value)
                else:
                    results.append(f"No valid {metric} data for sector {entity_name} in {sheet_name}.")

        if not results or not values:
            return f"No valid data found for {metric} across specified entities.", None

        # Generate comparison chart
        fig = go.Figure(data=[
            go.Bar(x=labels, y=values, marker_color='#36A2EB', name=metric)
        ])
        fig.update_layout(
            title=f'Comparison of {metric} Across Entities',
            xaxis_title='Entity',
            yaxis_title=metric,
            template='plotly_white'
        )

        result = f"Comparison of {metric}:\n" + "\n".join(results)
        logging.debug(f"Mixed entity comparison result: {result}")
        return result, fig

    def _handle_range_query(self, display_metric, filter_metric, min_value, max_value):
        """Handle queries for companies within a metric range."""
        sheet_name = detect_multirow_sheet(self.wb)
        display_columns = [col for sheet, col, _ in self._find_column_across_sheets(display_metric) if sheet == sheet_name]
        filter_columns = [col for sheet, col, _ in self._find_column_across_sheets(filter_metric) if sheet == sheet_name]
        if not display_columns or not filter_columns:
            return f"One or both metrics ('{display_metric}', '{filter_metric}') not found in {sheet_name}.", None

        sheet = self.sheet_info[sheet_name]['sheet']
        primary_key_col = self.sheet_info[sheet_name]['primary_key_col']
        col_idx = openpyxl.utils.column_index_from_string(primary_key_col) - 1
        data_rows = self.sheet_info[sheet_name]['data_rows']
        results = []
        chart_data = []

        for row in data_rows:
            code = sheet.cell(row=row, column=col_idx + 1).value
            if not code or not self._validate_entity(code.upper().strip(), is_company=True):
                continue
            filter_value = None
            for col in filter_columns:
                val = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                if isinstance(val, (int, float)):
                    filter_value = val
                    break
            if filter_value is None or not (min_value <= filter_value <= max_value):
                continue
            display_value = None
            for col in display_columns:
                val = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                if isinstance(val, (int, float)):
                    display_value = val
                    break
            if display_value is not None:
                results.append(f"{code}: {display_metric} = {display_value:.2f}")
                chart_data.append((code, display_value))

        if not results:
            return f"No companies found with {filter_metric} between {min_value} and {max_value} in {sheet_name}.", None

        result = f"Companies with {filter_metric} between {min_value} and {max_value}:\n" + "\n".join(results)
        fig = self._generate_company_chart(chart_data, display_metric)
        logging.debug(f"Range query result: {result}")
        return result, fig

    def _generate_company_chart(self, code, metrics):
        """Generate a bar chart for a company's metrics."""
        data = []
        labels = []
        for metric in metrics:
            matches = self._find_column_across_sheets(metric)
            for sheet_name, col, header in matches:
                row = self._find_company(sheet_name, code)
                if row:
                    sheet = self.sheet_info[sheet_name]['sheet']
                    value = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                    if isinstance(value, (int, float)):
                        labels.append(f"{metric} ({sheet_name})")
                        data.append(value)
        if not data:
            return None
        fig = go.Figure(data=[
            go.Bar(x=labels, y=data, marker_color='#36A2EB', name=f'Metrics for {code}')
        ])
        fig.update_layout(
            title=f'Metrics for {code}',
            xaxis_title='Metric',
            yaxis_title='Value',
            template='plotly_white'
        )
        return fig
    
    #### new section made on 17/10/2025 ####

    def _generate_sector_chart(self, sector, metric, company_data):
        """Generate a bar chart for a sector's metric with all companies."""
        if not company_data:
            return None
        
        # Sort companies by value (highest to lowest)
        company_data.sort(key=lambda x: x[1], reverse=True)
        
        # Limit to top 20 companies for better visualization
        if len(company_data) > 20:
            display_data = company_data[:20]
        else:
            display_data = company_data
        
        labels, values = zip(*display_data)
        
        fig = go.Figure(data=[
            go.Bar(x=labels, y=values, marker_color='#36A2EB', name=f'{metric}')
        ])
        
        fig.update_layout(
            title=f'{metric} by Company in {sector}',
            xaxis_title='Company',
            yaxis_title=metric,
            template='plotly_white',
            xaxis_tickangle=-45
        )
        
        return fig

    def _generate_best_stock_chart(self, sorted_scores, criteria):
        """Generate a bar chart for top stocks."""
        if not sorted_scores:
            return None
        labels, values = zip(*sorted_scores)
        fig = go.Figure(data=[
            go.Bar(x=labels, y=values, marker_color='#4BC0C0', name=f'Score ({criteria})')
        ])
        fig.update_layout(
            title=f'Top Stocks by {criteria}',
            xaxis_title='Company',
            yaxis_title='Score',
            template='plotly_white'
        )
        return fig
    
    #### New section done on 17/10/2025

    def _generate_sector_comparison_chart_all(self, sector_data, metric):
        """Generate a bar chart comparing all sectors for a metric."""
        if not sector_data:
            return None
        
        labels, values = zip(*sector_data)
        
        fig = go.Figure(data=[
            go.Bar(x=labels, y=values, marker_color='#9966FF', name=f'Average {metric}')
        ])
        
        fig.update_layout(
            title=f'Average {metric} by Sector',
            xaxis_title='Sector',
            yaxis_title=f'Average {metric}',
            template='plotly_white',
            xaxis_tickangle=-45
        )
        
        return fig

    def _generate_best_metric_chart(self, sorted_values, metric, direction):
        """Generate a bar chart for best metric values."""
        if not sorted_values:
            return None
        labels, values = zip(*sorted_values)
        fig = go.Figure(data=[
            go.Bar(x=labels, y=values, marker_color='#FFCE56', name=f'{metric} ({direction})')
        ])
        fig.update_layout(
            title=f'Top Companies by {metric} ({direction.upper()})',
            xaxis_title='Company',
            yaxis_title=metric,
            template='plotly_white'
        )
        return fig

    def _generate_multi_sheet_chart(self, code, metric_sheet_pairs):
        """Generate a bar chart for multi-sheet query."""
        data = []
        labels = []
        for metric, sheet_name in metric_sheet_pairs:
            matches = [(s, c, h) for s, c, h in self._find_column_across_sheets(metric) if s.lower() == sheet_name.lower()]
            for sheet_name, col, header in matches:
                row = self._find_company(sheet_name, code)
                if row:
                    sheet = self.sheet_info[sheet_name]['sheet']
                    value = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col)).value
                    if isinstance(value, (int, float)):
                        labels.append(f"{metric} ({sheet_name})")
                        data.append(value)
        if not data:
            return None
        fig = go.Figure(data=[
            go.Bar(x=labels, y=data, marker_color='#9966FF', name=f'Metrics for {code}')
        ])
        fig.update_layout(
            title=f'Metrics for {code} Across Sheets',
            xaxis_title='Metric',
            yaxis_title='Value',
            template='plotly_white'
        )
        return fig

    def _validate_entity(self, entity, is_company=True):
        """Validate if an entity is a company or sector."""
        entity = entity.upper().strip()
        for sheet_name in self.sheet_info:
            if is_company:
                if self._find_company(sheet_name, entity):
                    return True
            else:
                if self._get_sector_rows(sheet_name, entity):
                    return True
        return False

    def process_structured_query(self, query_dict):
        """Process a structured query from LangChain."""
        query_type = query_dict.get('type')
        if query_type == 'company':
            company = query_dict.get('company', '').upper().strip()
            if not self._validate_entity(company, is_company=True):
                return f"Company '{company}' not found in any sheet.", None
            return self._handle_company_multi_metric(company, query_dict.get('metrics', []))
        elif query_type == 'sector':
            sector = query_dict.get('sector', '').upper().strip()
            if not self._validate_entity(sector, is_company=False):
                return f"Sector '{sector}' not found in any sheet.", None
            return self._handle_sector_metric(sector, query_dict.get('metric', ''))
        elif query_type == 'general':
            return self._handle_general_metric(query_dict.get('metric', ''))
        elif query_type == 'definition':
            return self._define_header(query_dict.get('metric', ''))
        elif query_type == 'best_stock':
            return self._find_best_stock(query_dict.get('criteria'))
        elif query_type == 'best_sector':
            return self._find_best_sector(query_dict.get('criteria'))
        elif query_type == 'best_metric':
            return self._find_best_metric_value(query_dict.get('metric', ''))
        elif query_type == 'compare_stocks':
            stocks = [s.upper().strip() for s in query_dict.get('stocks', [])]
            invalid_stocks = [s for s in stocks if not self._validate_entity(s, is_company=True)]
            if invalid_stocks:
                return f"Companies {', '.join(invalid_stocks)} not found.", None
            metric = query_dict.get('metric', 'P/E')
            return self._compare_stocks(stocks, metric=metric)
        elif query_type == 'multi_sheet':
            company = query_dict.get('company', '').upper().strip()
            if not self._validate_entity(company, is_company=True):
                return f"Company '{company}' not found in any sheet.", None
            return self._handle_multi_sheet_query(company, query_dict.get('metric_sheet_pairs', []))
        elif query_type == 'compare_mixed':
            entities = query_dict.get('entities', [])
            invalid_entities = []
            for entity in entities:
                entity_name = entity.get('name', '').upper().strip()
                entity_type = entity.get('type')
                if entity_type == 'company' and not self._validate_entity(entity_name, is_company=True):
                    invalid_entities.append(entity_name)
                elif entity_type == 'sector' and not self._validate_entity(entity_name, is_company=False):
                    invalid_entities.append(entity_name)
            if invalid_entities:
                return f"Entities {', '.join(invalid_entities)} not found.", None
            metric = query_dict.get('metric', 'P/E')
            return self._compare_mixed_entities(entities, metric)
        return "Invalid query structure.", None

    def process_query(self, query):
        """Process a natural language query."""
        query = query.strip()
        normalized_query = query.lower()
        logging.info(f"Processing query: {query}")

        # Regex patterns
        multi_sheet_pattern = r'(.+?)\s+from\s+([\w\s\d]+)\s+and\s+(.+?)\s+from\s+([\w\s\d]+)\s+for\s+([a-z0-9\.]+)'
        multi_metric_pattern = r'(.+?)\s+for\s+([a-z0-9]+)'
        sector_pattern = r'(.+?)\s+for\s+sector\s+([\w\s,&]+)'
        general_pattern = r'average\s+(.+)'
        define_pattern = r'(?:what is|define)\s+(.+)'
        best_stock_pattern = r'(?:best stock|which stock is best|top stocks?)(?:\s+by\s+(.+))?'
        best_sector_pattern = r'(?:best sector|which sector is best|top sectors?)(?:\s+by\s+(.+))?'
        lowest_metric_pattern = r'lowest\s+([a-z0-9\.\/]+)'
        highest_metric_pattern = r'highest\s+([a-z0-9\.\/]+)'
        # best_metric_pattern = r'best\s+([a-z0-9\.\/]+)'
        best_metric_pattern = r'(?:best|lowest|highest)\s+([a-z0-9\.\/]+)'
        compare_stocks_pattern = r'(?:which stocks are best|compare stocks)\s+((?:[a-z0-9]+(?:\s*,\s*[a-z0-9]+)*))\s*(?:by\s+(.+))?'
        compare_mixed_pattern = r'([a-z0-9]+)\s+vs\s+([a-z0-9]+)\s+vs\s+sector\s+([\w\s,&]+)\s+by\s+(.+)'
        compare_sectors_pattern = r'sector\s+([\w\s,&]+)\s+vs\s+sector\s+([\w\s,&]+)\s+by\s+(.+)'
        range_pattern = r'show\s+(.+?)\s+where\s+(.+?)\s+between\s+([\d\.]+)\s+and\s+([\d\.]+)'
        stock_vs_sector_pattern = r'([a-z0-9]+)\s+vs\s+sector\s+([\w\s,&]+)\s+by\s+(.+)'
        sector_vs_sector_pattern = r'sector\s+([\w\s,&]+)\s+vs\s+sector\s+([\w\s,&]+)\s+by\s+(.+)'

        # Check for "lowest" first (most specific)
        lowest_match = re.search(lowest_metric_pattern, normalized_query)
        if lowest_match:
            metric = self._normalize_metric(lowest_match.group(1).strip())
            return self._find_lowest_metric_value(metric)
        
        # Check for "highest" second
        highest_match = re.search(highest_metric_pattern, normalized_query)
        if highest_match:
            metric = self._normalize_metric(highest_match.group(1).strip())
            return self._find_highest_metric_value(metric)

        # Handle range query
        range_match = re.search(range_pattern, normalized_query)
        if range_match:
            display_metric, filter_metric, min_value, max_value = range_match.groups()
            display_metric = self._normalize_metric(display_metric.strip())
            filter_metric = self._normalize_metric(filter_metric.strip())
            try:
                min_value, max_value = float(min_value), float(max_value)
                return self._handle_range_query(display_metric, filter_metric, min_value, max_value)
            except ValueError:
                return "Error: Invalid range values. Please enter numeric values.", None

        # Handle compare_sectors query
        compare_sectors_match = re.search(compare_sectors_pattern, normalized_query)
        if compare_sectors_match:
            sector1, sector2, metric = compare_sectors_match.groups()
            entities = [
                {'name': sector1.upper().strip(), 'type': 'sector'},
                {'name': sector2.upper().strip(), 'type': 'sector'}
            ]
            invalid_entities = []
            for entity in entities:
                if not self._validate_entity(entity['name'], is_company=False):
                    invalid_entities.append(entity['name'])
            if invalid_entities:
                return f"Sectors {', '.join(invalid_entities)} not found.", None
            metric = self._normalize_metric(metric.strip())
            return self._compare_mixed_entities(entities, metric)

        # Handle compare_mixed query
        compare_mixed_match = re.search(compare_mixed_pattern, normalized_query)
        if compare_mixed_match:
            stock1, stock2, sector, metric = compare_mixed_match.groups()
            entities = [
                {'name': stock1.upper().strip(), 'type': 'company'},
                {'name': stock2.upper().strip(), 'type': 'company'},
                {'name': sector.upper().strip(), 'type': 'sector'}
            ]
            invalid_entities = []
            for entity in entities:
                if not self._validate_entity(entity['name'], is_company=(entity['type'] == 'company')):
                    invalid_entities.append(entity['name'])
            if invalid_entities:
                return f"Entities {', '.join(invalid_entities)} not found.", None
            metric = self._normalize_metric(metric.strip())
            return self._compare_mixed_entities(entities, metric)

        # Handle stock vs sector query (NEW)
        stock_vs_sector_match = re.search(stock_vs_sector_pattern, normalized_query)
        if stock_vs_sector_match:
            stock, sector, metric = stock_vs_sector_match.groups()
            entities = [
                {'name': stock.upper().strip(), 'type': 'company'},
                {'name': sector.upper().strip(), 'type': 'sector'}
            ]
            invalid_entities = []
            for entity in entities:
                if not self._validate_entity(entity['name'], is_company=(entity['type'] == 'company')):
                    invalid_entities.append(entity['name'])
            if invalid_entities:
                return f"Entities {', '.join(invalid_entities)} not found.", None
            metric = self._normalize_metric(metric.strip())
            return self._compare_mixed_entities(entities, metric)

        # Handle sector vs sector query (NEW)
        sector_vs_sector_match = re.search(sector_vs_sector_pattern, normalized_query)
        if sector_vs_sector_match:
            sector1, sector2, metric = sector_vs_sector_match.groups()
            entities = [
                {'name': sector1.upper().strip(), 'type': 'sector'},
                {'name': sector2.upper().strip(), 'type': 'sector'}
            ]
            invalid_entities = []
            for entity in entities:
                if not self._validate_entity(entity['name'], is_company=False):
                    invalid_entities.append(entity['name'])
            if invalid_entities:
                return f"Sectors {', '.join(invalid_entities)} not found.", None
            metric = self._normalize_metric(metric.strip())
            return self._compare_mixed_entities(entities, metric)
        
        #### Rest of the old code ####

        compare_stocks_match = re.search(compare_stocks_pattern, normalized_query)
        if compare_stocks_match:
            stocks_str, metric = compare_stocks_match.groups()
            
            # Parse stock codes while preserving extensions
            stocks = []
            for part in re.split(r',\s*', stocks_str.strip()):
                part = part.strip().upper()
                if '.' in part and len(part.split('.')[-1]) <= 2:  # Valid extension
                    stocks.append(part)
                else:
                    # For codes without extensions, ensure we don't accidentally match extensions
                    stocks.append(part.split('.')[0])
            
            metric = self._normalize_metric(metric.strip() if metric else 'P/E')
            return self._compare_stocks(stocks, metric=metric)

        # Handle sector query        
        
        # Handle sector query updated on 17/10/2025

        sector_match = re.search(sector_pattern, normalized_query)
        if sector_match:
            metric, sector = sector_match.groups()
            metric = self._normalize_metric(metric.strip())
            sector = sector.upper().strip()
            
            # Handle "all sectors" case
            if sector == "ALL SECTORS":
                return self._handle_sector_metric("all sectors", metric)
            
            if not self._validate_entity(sector, is_company=False):
                return f"Sector '{sector}' not found in any sheet.", None
            return self._handle_sector_metric(sector, metric)

        # Handle multi-sheet query
        multi_sheet_match = re.search(multi_sheet_pattern, normalized_query)
        if multi_sheet_match:
            metric1, sheet1, metric2, sheet2, code = multi_sheet_match.groups()
            metric1 = self._normalize_metric(metric1.strip())
            metric2 = self._normalize_metric(metric2.strip())
            if not self._validate_entity(code, is_company=True):
                return f"Company '{code}' not found in any sheet.", None
            return self._handle_multi_sheet_query(code, [(metric1, sheet1), (metric2, sheet2)])

        # Handle multi-metric query (before single metric)
        if ' and ' in normalized_query and ' for ' in normalized_query:
            match = re.search(multi_metric_pattern, normalized_query)
            if match:
                metrics_str, code = match.groups()
                metrics = [self._normalize_metric(m.strip()) for m in metrics_str.split(' and ')]
                if not self._validate_entity(code, is_company=True):
                    return f"Company '{code}' not found in any sheet.", None
                return self._handle_company_multi_metric(code, metrics)

        # Handle single-metric query
        metric_match = re.search(multi_metric_pattern, normalized_query)
        if metric_match:
            metric, code = metric_match.groups()
            metric = self._normalize_metric(metric.strip())
            if not self._validate_entity(code, is_company=True):
                return f"Company '{code}' not found in any sheet.", None
            return self._handle_company_multi_metric(code, [metric])

        # Handle other query types
        define_match = re.search(define_pattern, normalized_query)
        if define_match:
            metric = self._normalize_metric(define_match.group(1).strip())
            return self._define_header(metric)

        best_stock_match = re.search(best_stock_pattern, normalized_query)
        if best_stock_match:
            criteria = self._normalize_metric(best_stock_match.group(1).strip()) if best_stock_match.group(1) else None
            return self._find_best_stock(criteria)

        best_sector_match = re.search(best_sector_pattern, normalized_query)
        if best_sector_match:
            criteria = self._normalize_metric(best_stock_match.group(1).strip()) if best_sector_match.group(1) else None
            return self._find_best_sector(criteria)

        best_metric_match = re.search(best_metric_pattern, normalized_query)
        if best_metric_match:
            metric = self._normalize_metric(best_metric_match.group(1).strip())
            return self._find_best_metric_value(metric)

        general_match = re.search(general_pattern, normalized_query)
        if general_match:
            metric = self._normalize_metric(general_match.group(1).strip())
            return self._handle_general_metric(metric)

        logging.debug("Query not understood")
        return "Sorry, I couldn't understand your query. Try 'P/E for ALLI', 'P/E for sector FOOD, BEVERAGE & TOBACCO', 'average P/E', 'what is P/E', 'best stock by P/E', 'best sector by P/E', or 'ALLI vs BOC vs sector BANKS by Revenue 3M'.", None